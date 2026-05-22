# -*- coding: utf-8 -*-
"""
脚本07：出图 + 报告 —— 6张专题地图 + Excel 统计报告
"""
import os, sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
os.environ['PROJ_LIB'] = os.path.join(os.path.dirname(os.path.abspath(sys.executable)), 'Lib', 'site-packages', 'rasterio', 'proj_data')

import numpy as np
import rasterio
from rasterio.plot import plotting_extent
import geopandas as gpd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
from matplotlib.patches import Patch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RESULT = os.path.join(BASE_DIR, '成果数据')
DISPLAY = os.path.join(BASE_DIR, '成果展示')
REPORT_DIR = os.path.join(BASE_DIR, '分析报告')
os.makedirs(DISPLAY, exist_ok=True)
os.makedirs(REPORT_DIR, exist_ok=True)

print('=' * 60)
print('  脚本07: 出图 + 报告')
print('=' * 60)

# 全局绘图设置
plt.rcParams['font.sans-serif'] = ['Microsoft YaHei', 'SimHei', 'DejaVu Sans']
plt.rcParams['axes.unicode_minus'] = False

# 读取已重投影到 EPSG:4527 的边界 (用 ogr2ogr 预转换, 规避 pyproj/PROJ 冲突)
boundary = gpd.read_file(os.path.join(BASE_DIR, '原始数据',
                                       'hangzhou_boundary_4527.geojson'))

# 预计算全局 extent (所有栅格同一尺寸同一投影)
with rasterio.open(os.path.join(RESULT, 'slope.tif')) as tmp:
    GLOBAL_EXTENT = plotting_extent(tmp)


def read_raster(path):
    """读取栅格数据 + extent"""
    with rasterio.open(path) as src:
        data = src.read(1)
        extent = plotting_extent(src)
    return data, extent


def make_map(data, extent, title, out_path, cmap='terrain',
             vmin=None, vmax=None, cbar_label='', discrete=False,
             nodata=None):
    """通用制图函数"""
    fig, ax = plt.subplots(figsize=(12, 10), dpi=150)

    if nodata is not None:
        data_plot = np.where(data == nodata, np.nan, data)
    else:
        data_plot = data

    if discrete:
        n_class = int(np.nanmax(data_plot)) + 1
        colors = plt.cm.get_cmap(cmap, n_class)
        im = ax.imshow(data_plot, cmap=colors, vmin=0, vmax=n_class,
                        extent=extent)
    else:
        im = ax.imshow(data_plot, cmap=cmap, vmin=vmin, vmax=vmax,
                        extent=extent)

    boundary.boundary.plot(ax=ax, color='black', linewidth=0.5, aspect=None)
    cbar = plt.colorbar(im, ax=ax, shrink=0.7, label=cbar_label)
    ax.set_title(title, fontsize=14)
    ax.set_xlabel('CGCS2000 Easting (m)')
    ax.set_ylabel('CGCS2000 Northing (m)')
    fig.tight_layout()
    fig.savefig(out_path, dpi=150, bbox_inches='tight')
    plt.close(fig)


# ============================================================
# 1a. 高程 + 山体阴影叠加图
# ============================================================
print('  [1/6] 高程渲染 + 山体阴影...')
dem, dem_extent = read_raster(os.path.join(BASE_DIR, '处理中', 'dem_filled.tif'))
hs, _ = read_raster(os.path.join(RESULT, 'hillshade.tif'))

fig, ax = plt.subplots(figsize=(14, 11), dpi=150)
dem_plot = np.where(dem > -9000, dem, np.nan)
im = ax.imshow(dem_plot, cmap='terrain', extent=dem_extent)
# 叠加山体阴影
hs_plot = np.where(hs != 255, hs, np.nan)
ax.imshow(hs_plot, cmap='gray', alpha=0.4, extent=dem_extent)
boundary.boundary.plot(ax=ax, color='white', linewidth=0.8, aspect=None)
cbar = plt.colorbar(im, ax=ax, shrink=0.7, label='Elevation (m)')
ax.set_title('Hangzhou DEM + Hillshade', fontsize=15)
ax.set_xlabel('CGCS2000 Easting (m)')
ax.set_ylabel('CGCS2000 Northing (m)')
fig.tight_layout()
fig.savefig(os.path.join(DISPLAY, '01_dem_hillshade.png'), dpi=150)
plt.close(fig)

# ============================================================
# 1b. 坡度图
# ============================================================
print('  [2/6] 坡度...')
slope, extent = read_raster(os.path.join(RESULT, 'slope.tif'))
slope_plot = np.where(slope > -9000, slope, np.nan)
make_map(slope_plot, extent, 'Slope',
         os.path.join(DISPLAY, '02_slope.png'),
         cmap='YlOrRd', vmax=np.nanpercentile(slope_plot, 98),
         cbar_label='Slope (degrees)')

# ============================================================
# 1c. 坡向图
# ============================================================
print('  [3/6] 坡向...')
aspect, extent = read_raster(os.path.join(RESULT, 'aspect.tif'))
aspect_plot = np.where(aspect >= 0, aspect, np.nan)
make_map(aspect_plot, extent, 'Aspect',
         os.path.join(DISPLAY, '03_aspect.png'),
         cmap='twilight', vmin=0, vmax=360, cbar_label='Aspect (deg)')

# ============================================================
# 1d. 地形起伏度
# ============================================================
print('  [4/6] 地形起伏度...')
tri, extent = read_raster(os.path.join(RESULT, 'tri.tif'))
tri_plot = np.where(tri > -9000, tri, np.nan)
make_map(tri_plot, extent, 'Terrain Ruggedness Index (TRI)',
         os.path.join(DISPLAY, '04_tri.png'),
         cmap='plasma', vmax=np.nanpercentile(tri_plot, 95),
         cbar_label='TRI')

# ============================================================
# 1e. 高程分级
# ============================================================
print('  [5/6] 高程分级...')
elev_z, extent = read_raster(os.path.join(RESULT, 'elevation_zones.tif'))
zone_colors = ['#f7fcf5', '#e5f5e0', '#a1dab4', '#41b6c4', '#2c7fb8', '#253494']
zone_labels = ['0-50m', '50-100m', '100-200m', '200-500m', '500-1000m', '>1000m']

fig, ax = plt.subplots(figsize=(14, 11), dpi=150)
cmap = mcolors.ListedColormap(zone_colors)
elev_plot = np.where(elev_z != 255, elev_z, np.nan)
im = ax.imshow(elev_plot, cmap=cmap, vmin=1, vmax=6, extent=extent,
                interpolation='nearest')
boundary.boundary.plot(ax=ax, color='black', linewidth=0.5, aspect=None)
ax.set_title('Elevation Zones', fontsize=15)
legend_patches = [Patch(color=zone_colors[i], label=zone_labels[i]) for i in range(6)]
ax.legend(handles=legend_patches, loc='lower right', fontsize=9)
ax.set_xlabel('CGCS2000 Easting (m)')
ax.set_ylabel('CGCS2000 Northing (m)')
fig.tight_layout()
fig.savefig(os.path.join(DISPLAY, '05_elevation_zones.png'), dpi=150)
plt.close(fig)

# ============================================================
# 1f. 坡度分级
# ============================================================
print('  [6/6] 坡度分级...')
slope_z, extent = read_raster(os.path.join(RESULT, 'slope_zones.tif'))
s_colors = ['#f7fcf5', '#c7e9c0', '#78c679', '#238443', '#feb24c', '#bd0026']
s_labels = ['0-2', '2-5', '5-15', '15-25', '25-45', '>45']

fig, ax = plt.subplots(figsize=(14, 11), dpi=150)
cmap = mcolors.ListedColormap(s_colors)
sl_plot = np.where(slope_z != 255, slope_z, np.nan)
im = ax.imshow(sl_plot, cmap=cmap, vmin=1, vmax=6, extent=extent,
                interpolation='nearest')
boundary.boundary.plot(ax=ax, color='black', linewidth=0.5, aspect=None)
ax.set_title('Slope Zones (degrees)', fontsize=15)
legend_patches = [Patch(color=s_colors[i], label=f'{s_labels[i]} deg') for i in range(6)]
ax.legend(handles=legend_patches, loc='lower right', fontsize=9)
ax.set_xlabel('CGCS2000 Easting (m)')
ax.set_ylabel('CGCS2000 Northing (m)')
fig.tight_layout()
fig.savefig(os.path.join(DISPLAY, '06_slope_zones.png'), dpi=150)
plt.close(fig)

sizes = [os.path.getsize(os.path.join(DISPLAY, f)) / 1024
         for f in os.listdir(DISPLAY) if f.endswith('.png')]
print(f'  6 张图生成完成 ({sum(sizes):.0f} KB)')

# ============================================================
# 2. Excel 报告
# ============================================================
print(f'\n[2] 生成 Excel 报告...')

wb = Workbook()
header_font = Font(name='Microsoft YaHei', size=12, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center')
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


# Sheet 1: 总体概览
ws1 = wb.active
ws1.title = '项目概览'
overview = [
    ['项目名称', '基于DEM的杭州市地形分析与栅格数据标准化处理'],
    ['数据源', 'Copernicus DEM GLO-30 (ESA, 30m)'],
    ['研究区', '杭州市 (浙江省)'],
    ['坐标系统', 'CGCS2000 / 3-degree Gauss-Kruger CM 120E (EPSG:4527)'],
    ['空间分辨率', '30m'],
    ['核心方法', 'Horn 1981坡度坡向 + Riley TRI + Weiss TPI + 高程坡度分级'],
    ['技术栈', 'Python + rasterio + GDAL + PostGIS + GeoServer'],
    ['输出', '9个栅格产品 + 等高线 + 6张专题图 + Excel报告 + WMS/WCS服务'],
]
for r, row in enumerate(overview, 1):
    for c, val in enumerate(row, 1):
        ws1.cell(row=r, column=c, value=val)
style_header(ws1, 1, 2)
ws1.column_dimensions['A'].width = 18
ws1.column_dimensions['B'].width = 55

# Sheet 2: 统计摘要
ws2 = wb.create_sheet('统计摘要')
dem_v = dem[dem > -9000]
slope_v = slope[slope > -9000]
aspect_v = aspect[aspect >= 0]
tri_v = tri[tri > -9000]

stats_data = [
    ['指标', '最小值', '最大值', '均值', '中位数', '标准差'],
    ['高程(m)', f'{dem_v.min():.0f}', f'{dem_v.max():.0f}',
     f'{dem_v.mean():.0f}', f'{np.median(dem_v):.0f}', f'{dem_v.std():.0f}'],
    ['坡度(deg)', f'{np.nanmin(slope_v):.1f}', f'{np.nanmax(slope_v):.1f}',
     f'{np.nanmean(slope_v):.1f}', f'{np.nanmedian(slope_v):.1f}',
     f'{np.nanstd(slope_v):.1f}'],
    ['坡向(deg)', '0', '360',
     f'{np.nanmean(aspect_v):.0f}', f'{np.nanmedian(aspect_v):.0f}',
     f'{np.nanstd(aspect_v):.0f}'],
    ['起伏度TRI', f'{np.nanmin(tri_v):.3f}', f'{np.nanmax(tri_v):.3f}',
     f'{np.nanmean(tri_v):.3f}', f'{np.nanmedian(tri_v):.3f}',
     f'{np.nanstd(tri_v):.3f}'],
]
for r, row in enumerate(stats_data, 1):
    for c, val in enumerate(row, 1):
        ws2.cell(row=r, column=c, value=val)
style_header(ws2, 1, 6)
for c in range(1, 7):
    ws2.column_dimensions[chr(64 + c)].width = 14

# Sheet 3: 地形分级统计
ws3 = wb.create_sheet('地形分级')
elev_zone_names = ['0-50m平原', '50-100m台地', '100-200m丘陵',
                   '200-500m低山', '500-1000m中山', '>1000m高中山']
elev_counts = [(elev_z == i + 1).sum() for i in range(6)]
elev_total = sum(elev_counts)
elev_rows = [['高程分级', '像素数', '占比(%)']]
for i, (name, cnt) in enumerate(zip(elev_zone_names, elev_counts)):
    elev_rows.append([name, cnt, f'{cnt/elev_total*100:.1f}'])

slope_zone_names = ['0-2deg平坦', '2-5deg微坡', '5-15deg缓坡',
                    '15-25deg中坡', '25-45deg陡坡', '>45deg极陡坡']
slope_counts = [(slope_z == i + 1).sum() for i in range(6)]
slope_total = sum(slope_counts)

for r, row in enumerate(elev_rows, 1):
    for c, val in enumerate(row, 1):
        ws3.cell(row=r, column=c, value=val)
style_header(ws3, 1, 3)

ws3.cell(row=1, column=5, value='坡度分级')
ws3.cell(row=1, column=6, value='像素数')
ws3.cell(row=1, column=7, value='占比(%)')
ws3.cell(row=1, column=5).font = header_font
ws3.cell(row=1, column=5).fill = header_fill
ws3.cell(row=1, column=6).font = header_font
ws3.cell(row=1, column=6).fill = header_fill
ws3.cell(row=1, column=7).font = header_font
ws3.cell(row=1, column=7).fill = header_fill

for i, (name, cnt) in enumerate(zip(slope_zone_names, slope_counts)):
    ws3.cell(row=i + 2, column=5, value=name)
    ws3.cell(row=i + 2, column=6, value=cnt)
    ws3.cell(row=i + 2, column=7, value=f'{cnt/slope_total*100:.1f}')

for c in range(1, 8):
    ws3.column_dimensions[chr(64 + c)].width = 18

# Sheet 4: 方法说明
ws4 = wb.create_sheet('方法说明')
methods = [
    ['步骤', '方法说明'],
    ['数据获取', 'Copernicus DEM GLO-30 (ESA), AWS Open Data 在线下载, 6瓦片拼接'],
    ['预处理', '重投影到 CGCS2000(EPSG:4527), 30m分辨率, 填洼处理消除局部凹陷'],
    ['坡度坡向', 'Horn (1981) 3x3移动窗口算法, 8邻域加权差分'],
    ['山体阴影', 'Hillshade: 光源方位角315deg (西北), 高度角45deg (制图惯例)'],
    ['地形起伏度', 'Riley et al. (1999) TRI, 3x3窗口 RMS'],
    ['高程分级', '1:100万地貌图规范 6级: 平原/台地/丘陵/低山/中山/高中山'],
    ['坡度分级', '水土保持规范 6级: 平坦(<2)/微坡(2-5)/缓坡(5-15)/中坡(15-25)/'],
    ['', '陡坡(25-45)/极陡坡(>45)'],
    ['地形位指数', 'Weiss (2001) TPI: 300m半径邻域均值偏差, 7级分类'],
    ['入库', 'raster2pgsql 栅格入库 + ogr2ogr 矢量入库, schema=terrain'],
    ['服务', 'GeoServer REST API 发布 WMS/WCS/WFS'],
]
for r, row in enumerate(methods, 1):
    for c, val in enumerate(row, 1):
        ws4.cell(row=r, column=c, value=val)
style_header(ws4, 1, 2)
ws4.column_dimensions['A'].width = 18
ws4.column_dimensions['B'].width = 65

xlsx_path = os.path.join(REPORT_DIR, 'DEM地形分析报告.xlsx')
wb.save(xlsx_path)
print(f'  [OK] Excel → {xlsx_path}')

# ============================================================
# 3. 完成
# ============================================================
print(f'\n{"=" * 60}')
print(f'  出图报告完成')
print(f'  专题图: {len([f for f in os.listdir(DISPLAY) if f.endswith(".png")])} 张')
import json
summary_json = {
    '研究区': '杭州市',
    '数据': 'Copernicus DEM GLO-30 30m',
    '高程范围(m)': f'{dem_v.min():.0f}-{dem_v.max():.0f}',
    '平均高程(m)': f'{dem_v.mean():.0f}',
    '平均坡度(deg)': f'{np.nanmean(slope_v):.1f}',
    '最大坡度(deg)': f'{np.nanmax(slope_v):.1f}',
    '平原占比(%)': f'{elev_counts[0]/elev_total*100:.1f}',
    '山区占比(>200m,%)': f'{sum(elev_counts[3:])/elev_total*100:.1f}',
    '平坦区(<2deg,%)': f'{slope_counts[0]/slope_total*100:.1f}',
    '陡坡区(>25deg,%)': f'{sum(slope_counts[4:])/slope_total*100:.1f}',
}
with open(os.path.join(RESULT, 'terrain_summary.json'), 'w', encoding='utf-8') as f:
    json.dump(summary_json, f, ensure_ascii=False, indent=2)

print(f'  [OK] terrain_summary.json')
print(f'{"=" * 60}')
